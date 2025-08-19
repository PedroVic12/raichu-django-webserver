import openpyxl
from openpyxl.styles import Font, Alignment,PatternFill,Border,Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import re
import pandas as pd
import os

class Palkia():
    def __init__(self, nome_arquivo="planilha.xlsx", sheet_name="Sheet1"):
        self.wb = openpyxl.Workbook()
        self.nome_arquivo = nome_arquivo
        self.number_of_pages = len(self.wb.worksheets)

        # Se o nome da aba for diferente do padrão "Sheet", renomeia a primeira aba
        if sheet_name != "Sheet1":
            self.wb.active.title = sheet_name


    def get_sheets(self):
        return self.wb.sheetnames

    # Colocando Dados
    def add_title(self, sheet_name, cell, title, font_size=12):
        """
        Adiciona um título em uma determinada célula.

        Parameters:
        - sheet_name: Nome da aba no Excel.
        - cell: Célula onde o título deve ser colocado.
        - title: Título a ser adicionado.
        - font_size: Tamanho da fonte do título.
        """
        ws = self.wb[sheet_name]
        ws[cell] = title
        ws[cell].font = Font(bold=True, size=font_size)
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        #ws[cell].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    def add_dataframe(self, sheet_name, df, start_row, start_col, color_option="verde_claro"):
        """
        Adiciona um DataFrame abaixo de uma célula especificada.

        Parameters:
        - sheet_name: Nome da aba no Excel.
        - df: DataFrame a ser adicionado.
        - start_row: Linha inicial para começar a adicionar os dados.
        - start_col: Coluna inicial para começar a adicionar os dados.
        - color_option: Opção de cor para o preenchimento das células. Pode ser 'verde_claro', 'azul_claro' ou 'laranja_claro'.
        """

        
        # Define opções de cores
        colors = {
            "verde_claro": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
            "azul_claro": PatternFill(start_color="A9D9F7", end_color="A9D9F7", fill_type="solid"),
            "laranja_claro": PatternFill(start_color="FED9CC", end_color="FED9CC", fill_type="solid")
        }

        # Define a borda
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        ws = self.wb[sheet_name]
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Se é a primeira linha (cabeçalho), aplique a fonte bold e tamanho 14
                if r_idx == start_row:
                    cell.font = Font(bold=True, size=14)
                
                # Aplica a cor de preenchimento escolhida
                cell.fill = colors[color_option]
                
                # Aplica a borda
                cell.border = thin_border

    #Utilidades
    def get_last_row(self, sheet_name, column):
        """
        Retorna o número da última linha com dados em uma coluna específica.

        Parameters:
        - sheet_name: Nome da aba no Excel.
        - column: Coluna para verificar (e.g., 'A', 'B', ...).

        Returns:
        - Número da última linha com dados na coluna especificada.
        """
        ws = self.wb[sheet_name]
        for row in range(ws.max_row, 0, -1):
            if ws[column + str(row)].value:
                return row
        return 0

    def merge_cells_range(self, sheet_name, cell_range, alignment=True):
        """
        Mescla um range de células e opcionalmente as alinha ao centro.

        Parameters:
        - sheet_name: Nome da aba no Excel.
        - cell_range: Range de células a serem mescladas (por exemplo, 'A1:B2').
        - alignment (optional): Se True, alinha o conteúdo das células ao centro.
        """
        ws = self.wb[sheet_name]
        start_cell = cell_range.split(':')[0]
        
        ws.merge_cells(cell_range)
        
        if alignment:
            cell = ws[start_cell]
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def clear_worksheet(self, sheet_name):
        if sheet_name in self.wb.sheetnames:
            std = self.wb[sheet_name]
            self.wb.remove(std)
        self.wb.create_sheet(title=sheet_name)


    #Formtação de células
    def customize_font(self, sheet_name, cell, size=None, color=None, bold=None, italic=None):
        """
        Personaliza a fonte de uma célula específica.

        Parameters:
        - sheet_name: Nome da aba no Excel.
        - cell: Célula a ser personalizada.
        - size: Tamanho da fonte. Default é None (mantém o tamanho atual).
        - color: Cor da fonte (formato hexadecimal). Default é None (mantém a cor atual).
        - bold: True para fonte em negrito, False para regular. Default é None (mantém o estilo atual).
        - italic: True para fonte em itálico, False para regular. Default é None (mantém o estilo atual).
        """
        ws = self.wb[sheet_name]
        current_font = ws[cell].font

        # Use os valores atuais se os novos valores não forem fornecidos
        if size is None:
            size = current_font.sz
        if color is None:
            color = current_font.color
        if bold is None:
            bold = current_font.bold
        if italic is None:
            italic = current_font.italic

        ws[cell].font = Font(size=size, color=color, bold=bold, italic=italic)

    def colorir_colunasTabelas(self, sheet_name, column_index):
        """
        Colors the background of specific rows based on a condition.

        Parameters:
        - sheet_name: Name of the Excel sheet.
        - column_index: Index of the column to apply the formatting (0-based index).
        """
        ws = self.wb[sheet_name]
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # Red fill color

        for row in ws.iter_rows(min_row=2, max_col=column_index + 1, max_row=ws.max_row):  # Start from row 2 and specified column
            if row[column_index].value in ("Ações de manutenção 1", "Medidas de segurança 1"):  # Check if the cell value matches the target values
                for cell in row:
                    cell.fill = red_fill

    # Salvando
    def organizar_dataframe(self, df, output_path):
        """
        Esta função pega um DataFrame como entrada e processa suas células para dividir certos valores em várias células, com base
        em quebras de linha ou outros delimitadores, e então salva o DataFrame processado em um arquivo Excel.
        """
        # Criando um novo workbook e worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Preenchendo o worksheet com os dados do DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Processando colunas específicas
                if cell.row == 1 and "Ações de manutenção" in str(cell.value):
                    # Esta coluna precisa de processamento especial nas próximas linhas
                    continue
                if "Ações de manutenção" in ws.cell(row=1, column=c_idx).value:
                    # Divide o valor usando regex para identificar padrões de números seguidos de parênteses
                    segments = re.findall(r'(\d+\) .+?\.)', str(value))
                    for sub_idx, segment in enumerate(segments, 0):
                        ws.cell(row=r_idx + sub_idx, column=c_idx, value=segment)

        # Salvar o workbook processado
        wb.save(output_path)
        print('Salvo em ', output_path)



    def save(self):
        """
        Salva as alterações feitas no arquivo Excel.

        Parameters:
        - filename: Nome do arquivo onde as alterações devem ser salvas.
        - sheet_name: Nome da aba no Excel.
        """

        

        self.wb.save(self.nome_arquivo)
        print('Salvo em ', self.nome_arquivo)

    def format_columns(self, sheet_name, start_col, end_col=None, width=60):
        """
        Formata as colunas especificadas no intervalo.

        Parameters:
        - sheet_name: Nome da aba no Excel.
        - start_col: Letra da coluna inicial (por exemplo, 'A').
        - end_col (optional): Letra da coluna final (por exemplo, 'F'). Se não for fornecido, formatará apenas a start_col.
        - width (optional): Largura para definir para as colunas. Default é 60.
        """
        ws = self.wb[sheet_name]
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Se end_col não for fornecido, defina-o como start_col
        if not end_col:
            end_col = start_col
        
        for col_num in range(ord(start_col) - ord('A') + 1, ord(end_col) - ord('A') + 2):  # Convertendo letras para números de colunas
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width
            for row in ws[col_letter]:
                row.alignment = alignment

    def save_to_sheet(self, dataframe):
        """
        Salva o dataframe na aba especificada.

        Args:
        - dataframe (pd.DataFrame): DataFrame que será salvo.
        """
        with pd.ExcelWriter(self.nome_arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            writer.book = openpyxl.load_workbook(self.nome_arquivo)  # Carregar o arquivo Excel existente
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            dataframe.to_excel(writer, sheet_name=self.sheet_name, index=False)
            writer.save()
            print(f'Dados salvos na aba {self.sheet_name} do arquivo {self.nome_arquivo}')

    def merge_excel_files(self,directory, output_filename):
        """
        This function merges all Excel files in a given directory into a single Excel file.
        Each file is represented in a separate sheet.
        
        Parameters:
        - directory: The directory containing the Excel files to merge.
        - output_filename: The name of the resulting merged Excel file.
        """
        
        # List all files in the given directory
        filenames = os.listdir(directory)
        
        # Filter the list of filenames to include only Excel files
        excel_files = [f for f in filenames if f.endswith('.xlsx') or f.endswith('.xls')]
        
        if not excel_files:  # If there are no Excel files, simply return
            print(f"No Excel files found in {directory}.")
            return
        
        # Create a new Excel Workbook
        merged_workbook = openpyxl.Workbook()
        
        # If there are Excel files to merge, create the first sheet before removing the default one
        merged_workbook.create_sheet('Sheet1')
        merged_workbook.remove(merged_workbook.active)
        
        with pd.ExcelWriter(output_filename, engine='openpyxl', mode='w') as writer:
            # Iterate over each Excel file and add its content to a new sheet in the merged workbook
            for excel_file in excel_files:
                # Load the Excel file
                filepath = os.path.join(directory, excel_file)
                
                for sheetname in pd.ExcelFile(filepath).sheet_names:
                    # Read the sheet content into a DataFrame
                    data = pd.read_excel(filepath, sheet_name=sheetname)
                    
                    # Use apenas o nome da aba ao adicionar ao arquivo consolidado
                    data.to_excel(writer, sheet_name=sheetname, index=False)
        
        print(f"All Excel files in {directory} have been merged into {output_filename}.")

    def format_excel(self, filename):
        """
        Formata o arquivo Excel especificado. Esta função ajusta o tamanho das colunas para 60 e ativa a quebra de linha para todas as células.

        Parameters:
        - filename: Nome do arquivo Excel a ser formatado.
        """
        wb = openpyxl.load_workbook(filename)

        for sheet in wb.worksheets:
            for column in sheet.columns:
                column_letter = get_column_letter(column[0].column)
                
                # Definir a largura da coluna para 60
                sheet.column_dimensions[column_letter].width = 30
                
                # Ativar a quebra de linha para todas as células na coluna
                for cell in column:
                    cell.alignment = Alignment(wrap_text=True)

        # Salvar as alterações no arquivo
        wb.save(filename)
        print('Salvo em ', filename)


